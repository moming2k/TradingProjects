#!/usr/bin/env python
# -*- coding: utf-8 -*-

# @Filename: sort_20161212_loop_result
# @Date: 2016-12-13
# @Author: Mark Wang
# @Email: wangyouan@gmial.com

import os

import openpyxl
from bs4 import BeautifulSoup

root_path = r'F:\My Documents\Project\QuestionFromProfWang\GlassDoor\20161212_loop_result'
result_path = os.path.join(root_path, 'output')
absda_path = os.path.join(root_path, 'loop_absda_xxxP_simple1k')
resid_path = os.path.join(root_path, 'loop_resid_std_xxxP_simple1k')

def read_excel_xml(path):
    file = open(path).read()
    soup = BeautifulSoup(file, 'xml')
    workbook = []
    for sheet in soup.findAll('Worksheet'):
        sheet_as_list = []
        for row in sheet.findAll('Row'):
            row_as_list = []
            for cell in row.findAll('Cell'):
                row_as_list.append(cell.Data.text)
            sheet_as_list.append(row_as_list)
        workbook.append(sheet_as_list)
    return workbook


if __name__ == '__main__':
    data_path = resid_path
    for file_name in os.listdir(data_path):
        if not file_name.endswith('_1.xls'):
            continue
        prefix = file_name[:-6]
        wb = openpyxl.Workbook()
        ws = wb.active
        col_index = 1
        print file_name
        for i in range(1, 18):
            file_path = os.path.join(data_path, '{}_{}.xls'.format(prefix, i))
            if not os.path.isfile(os.path.join(data_path, '{}_{}.xls'.format(prefix, i))):
                break
            xls_info = read_excel_xml(file_path)[0]
            max_length = max(map(len, xls_info))

            for row_index in  range(len(xls_info)):
                current_col_index = col_index
                row = xls_info[row_index]
                if col_index == 1:
                    for value in row:
                        ws.cell(row=row_index + 1, column=current_col_index).value = value
                        current_col_index += 1
                else:
                    for value in row[1:]:
                        ws.cell(row=row_index + 1, column=current_col_index).value = value
                        current_col_index += 1

            col_index += max_length if col_index == 1 else max_length - 1
        wb.save(os.path.join(result_path, '{}.xlsx'.format(prefix)))
